#!/usr/bin/env python3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
HF = PatternFill('solid', fgColor='0D3330')
HFont = Font(name='Arial', bold=True, color='FFFFFF', size=11)
DF = Font(name='Arial', size=10)
TB = Border(left=Side('thin',color='CCCCCC'),right=Side('thin',color='CCCCCC'),top=Side('thin',color='CCCCCC'),bottom=Side('thin',color='CCCCCC'))
SF = PatternFill('solid', fgColor='D4EDDA')
AF = PatternFill('solid', fgColor='FFF3CD')
BF = PatternFill('solid', fgColor='D1ECF1')
CF = PatternFill('solid', fgColor='F8D7DA')
TIER_FILLS = {"S": SF, "A": AF, "B": BF, "C": CF}

def sh(ws, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font, cell.border = HF, HFont, TB
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def sd(ws, rows, cols):
    for r in range(2, rows+1):
        for c in range(1, cols+1):
            cell = ws.cell(row=r, column=c)
            cell.font, cell.border = DF, TB
            cell.alignment = Alignment(vertical='top', wrap_text=True)

# ===== SHEET 1: TOP 50 =====
ws1 = wb.active
ws1.title = "Top 50"
ws1.append(["Rank","Tier","School Name","Country","City/Region","# Schools","Key Person","Role","Email","Phone","Website","Score /40","Grade","Notes"])
data = [
[1,"S","Guidepost Montessori / CEG","USA + Asia","Brooklyn NY / Hong Kong","100+","Steve Xu","Founder & Global CEO","s***@guidepostmontessori.com","+852 9126 6211","guidepostmontessori.com",38,"B","Largest global network. ONE email = 100+ schools."],
[2,"S","Spring Education / LePort","USA","Campbell CA / West Chester PA","200+","CEO (Collins/Listman)","CEO","","484-947-2000","springeducationgroup.com",36,"B","Owns LePort, Stratford, Nobel, BASIS. AMI+AMS."],
[3,"S","Etonkids Educational Group","China","Beijing + 17 cities","60+","Vivien Wang","Chairman & CEO","lidoadmin@etonkids.com","8610 82357608","etonkids.com",35,"A","First ISO 9001 in China. Harvard/Kellogg. AMS."],
[4,"S","Montessori Academy Australia","Australia","NSW, VIC, ACT","70+","Corporate","","","1300 000 162","montessoriacademy.com.au",34,"C","96% Meeting/Exceeding National Quality Standard."],
[5,"S","MSB Beijing","China","Beijing","1","Alice Zhang","Chinese Dir","@msb.edu.cn","+86 10 6432 8228","msb.edu.cn",34,"B","Longest-running Montessori in China. 400 students."],
[6,"S","Brainy Bunch Int'l Islamic","Malaysia+5","Kuala Lumpur HQ","120","Fadzil Hashim / Efizah Rasali","CEO / Founder","bbisjohor@brainybunch.com","+60 10-271 9346","brainybunch.com",33,"B","World Leader Islamic Montessori. 6 countries."],
[7,"S","Montessori School of Tokyo","Japan","Tokyo","1","James Moore","Head of School","info@montessorijapan.com","03-5449-7067","montessorijapan.com",33,"A","First AMI accredited in Asia. Ages 2-15."],
[8,"S","Maria Montessori School","UK","London (3 sites)","3","Leadership team","","","+44 20 7435 3646","mariamontessori.org",32,"B","90%+ AMI teachers. Ages 2.5-16."],
[9,"A","QAIS Qingdao","China","Qingdao","1","Rosemary Gosse","Montessori Coord","Info@QingdaoAmerasia.org","","qingdaoamerasia.org",30,"A","First & only AMS accredited in Asia."],
[10,"A","MMI Singapore","Singapore","Singapore","Multiple","Corporate (1989)","","enquiry@modern-montessori.com","6220 8200","modern-montessori.com",30,"A","Franchisor since 1989. London origins."],
]
for r in data:
    ws1.append(r)
data2 = [
[11,"A","MABIS Bangkok","Thailand","Bangkok","1","Serene Jiratanan","Principal","info@montessoribkk.com","+66 87 828 2868","montessoribkk.com",30,"A","First non-US WASC Montessori. MIT principal."],
[12,"A","Peterson Schools","Mexico","Mexico City (4)","4","","","","","peterson.edu.mx",30,"B","1800 students. Oldest EN Montessori in Mexico."],
[13,"A","City Montessori School","India","Lucknow","21","","","","","cmseducation.org",30,"C","Guinness Record: 60,000 students."],
[14,"A","Nebula School Shanghai","China","Shanghai (5)","5","Campus Directors","","","","nebula website",30,"C","Bilingual Montessori. Credentialed teachers."],
[15,"A","Guidepost Shanghai","China","Shanghai","1","Part of GGE","","","","guidepostmontessori.com",30,"B","Mandarin + English. Global network."],
[16,"A","Fuji Kindergarten","Japan","Tokyo","1","Sekiichi Kato","Principal","","","fujikids.jp",30,"C","650 children. UNESCO featured."],
[17,"A","Sapientia Montessori","USA","Cedar Park TX","3","Kalika Sarathkumara","Admin","director@sapientiamontessori.com","512-260-2261","sapientiamontessori.com",28,"A","AMI Recognized. Oldest family N Austin."],
[18,"A","Montessori Friends Berlin","Germany","Berlin (2)","2","","","kita@montessori-friends.de","030-233264033","montessori-friends.de",28,"A","Bilingual EN/DE. Kinderhaus + IMS."],
[19,"A","Humberside Montessori","Canada","Toronto ON","1","","","","416-762-8888","humbersidemontessori.com",28,"B","AMI Recognized 2.5-12."],
[20,"A","La Villa Montessori","Canada","Mississauga ON","1","","","","905-822-2223","lvms.ca",28,"B","AMI Recognized all levels."],
[21,"A","Maria Montessori School","Canada","Toronto ON","1","","","","416-423-9123","mariamontessori.ca",28,"B","AMI Recognized all levels."],
[22,"A","Int'l Montessori Brussels","Belgium","Brussels (2)","2","","","","+32 2-721 2111","international-montessori.org",28,"B","Bilingual Montessori + IB World School."],
[23,"A","HD Qingdao Wanda","China","Qingdao","1","Ms. Li Suxiang","Principal","","","hdschools.org",28,"C","AMI 3-6 certified kindergarten."],
[24,"A","Montessori For Children","Singapore","Singapore","Multiple","Est. 1992","","","","montessori.edu.sg",28,"B","AMI standards since 1992."],
[25,"B","The Magic Years","India","Delhi","1","Monica Sagar","Director","monica.sagar@sns.edu.in","011-26140317","themagicyears.com",27,"A","Premium Delhi Montessori."],
]
for r in data2:
    ws1.append(r)
data3 = [
[26,"B","Montessori del Bosque","Mexico","Mexico City","1","Pioneers","","","","montessoridelbosque.com",27,"B","Pioneers of Montessori in all of Mexico."],
[27,"B","Greentree Montessori","Singapore","Singapore","Multiple","Est. 1990","","","","greentreemontessori.com",27,"B","35+ yrs. AMI teachers. Multiple campuses."],
[28,"B","Lodestar Montessori","Singapore","Singapore","Multiple","","","","","lodestarmontessori.com",27,"B","Pre-school through high school."],
[29,"B","Int'l Montessori Dubai","UAE","Dubai","1","","","","","montessorischooldubai.com",27,"B","KHDA accredited."],
[30,"B","Key Int'l School","Kenya","Nairobi","1","","","ims@montessori.co.ke","0727 810 350","keyinternationalschool.com",26,"A","East Africa's first int'l Montessori."],
[31,"B","Sweven Montessori","India","Mumbai","1","Mala Sharma","","school@swevenedu.com","9820155012","swevenmontessori.com",26,"A","Indian Montessori Foundation flagship."],
[32,"B","The Ardee School","India","Delhi","1","Sunpritt Dang","Coordinator","preprimary_coordinator_nfc@theardeeschool.com","18001027333","theardeeschool.com",26,"A","Large Delhi pre-primary."],
[33,"B","Redwood Montessori","UAE","Dubai+Abu Dhabi","Multiple","Kids First Group","","","+971-4-422-8581","",26,"B","Multiple UAE branches."],
[34,"B","Lions Gate Montessori","Canada","Vancouver BC","1","","","","604-677-1958","lionsgatemontessori.org",26,"B","AMI. Central Vancouver."],
[35,"B","Premier Montessori Acad","Canada","Richmond BC","1","","","","604-214-2927","premiermontessori.ca",26,"B","AMI Recognized 2 levels."],
[36,"B","Montessori Kinderhaus MIA","Germany","Munich","1","AMI affiliated","","","","montessori-ami-edu.de",26,"B","Bilingual EN/DE. AMI."],
[37,"B","Rose House Montessori","UK","London (3)","3","Founded 2005","","","","rosehousmontessori.com",26,"C","Ages 2-11."],
[38,"B","Creative Minds Montessori","USA","New York","1","Top AMI in NY","","","","Via AMI locator",26,"C","#1 AMI-ranked in New York."],
[39,"B","Centennial Montessori","USA","California","1","Top AMI in CA","","","","Via AMI locator",26,"C","Top AMI-ranked in California."],
[40,"B","Sydney Montessori School","Australia","Sydney","1","","","","","montessori.org.au",26,"C","Montessori Australia listed."],
[41,"C","Tara International","India","","1","Karma Y. Doma","Principal","principal@tarainternationalschool.com","7872970550","",25,"A","Direct principal email."],
[42,"C","IFS School","India","Bangalore","1","Ramswarup Gorontla","Admin","admin@ifs.school","080-26653777","ifs.school",25,"A","Bangalore AMI school."],
[43,"C","Brainy Child Montessori","Singapore","Orchard Rd","1","","","enquiry@brainychild.sg","+65 6733 7669","brainychildmontessori.sg",25,"A","Premium Orchard Rd location."],
[44,"C","IMC Bangkok","Thailand","Bangkok","1","","","","","imc.ac.th",25,"B","EN/Thai/Chinese. Ages 2-6."],
[45,"C","IMI Delft","Netherlands","Delft","1","Training","","","+31 642067136","imi-global.nl",25,"B","AMI training institute."],
[46,"C","Together We Montessori","UAE","Dubai","1","Founders","","","","togetherwemontessori.com",25,"C","18+ yrs. AMI + Cache L5."],
[47,"C","Rainbow Montessori","UK","London (2)","2","Linda Madden","Founder","","","rainbowmontessori.co.uk",25,"C","Est. 1982."],
[48,"C","Nairobi Montessori","Kenya","Nairobi","1","","","","","nairobimontessorischool.org",24,"B","Ages 15mo-12yrs."],
[49,"C","Kria Montessori House","India","","1","Prasannitha Rao","","hello@kriamontessorihouse.com","91777 45533","kriamontessorihouse.com",24,"A","Direct email."],
[50,"C","Diya Montessori","India","Pune","1","Sharan Sirur","","info@diyamontessori.in","9422772686","diyamontessori.in",24,"A","AMI-style. Direct email."],
]
for r in data3:
    ws1.append(r)

# Style Sheet 1
sh(ws1, 14)
sd(ws1, 51, 14)
for i, w in enumerate([6,5,35,18,22,10,25,20,35,20,28,10,7,45], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
for r in range(2, 52):
    tier = ws1.cell(row=r, column=2).value
    f = TIER_FILLS.get(tier)
    if f:
        for c in range(1, 15):
            ws1.cell(row=r, column=c).fill = f
ws1.freeze_panes = 'A2'
ws1.auto_filter.ref = "A1:N51"

# ===== SHEET 2: CHAINS =====
ws2 = wb.create_sheet("Chains")
ws2.append(["Chain","HQ","# Schools","Countries","Key Person","Role","Email","Phone","Website","Accreditation","Notes"])
chains = [
["Guidepost / CEG","Brooklyn NY / HK","100+","USA,HK,China,Bali+","Steve Xu","Global CEO","s***@guidepostmontessori.com","+852 9126 6211","guidepostmontessori.com","AMI","Largest global network. 10K+ families."],
["Spring Education/LePort","Campbell CA","200+","USA (19 states)","CEO","CEO","","484-947-2000","springeducationgroup.com","AMI+AMS","LePort,Stratford,Nobel,BASIS."],
["Etonkids","Beijing","60+","China+Sydney","Vivien Wang","Chairman&CEO","lidoadmin@etonkids.com","8610 82357608","etonkids.com","AMS/MACTE","Harvard/Kellogg. ISO 9001."],
["Brainy Bunch","KL, Malaysia","120","6 countries","Fadzil Hashim","CEO","bbisjohor@brainybunch.com","+60 10-271 9346","brainybunch.com","","Islamic Montessori leader."],
["Montessori Academy AU","Sydney","70+","Australia","Corporate","","","1300 000 162","montessoriacademy.com.au","","96% Meeting/Exceeding."],
["MKU","USA","12","USA","Franchise","","","","montessorikidsuniverse.com","","Fee $37.5K-$64.5K."],
["Hopscotch","NY/Toronto/Kyiv","3 cities","USA,CA,UA","","","","","hopscotchmontessori.com","","270 families, 30+ langs."],
]
for r in chains:
    ws2.append(r)
sh(ws2, 11)
sd(ws2, 8, 11)
for i, w in enumerate([30,22,10,25,20,15,35,20,30,12,45], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A2'

# ===== SHEET 3: TRAINING ORGS =====
ws3 = wb.create_sheet("Training & Orgs")
ws3.append(["Organization","Location","Contact","Website","Relevance"])
orgs = [
["AMI (Association Montessori Internationale)","Amsterdam, NL","info@montessori-ami.org, +31 20 6798932","montessori-ami.org","Global standards. Launching Global School Accreditation 2025-26."],
["AMI/USA","USA","amiusa.org","amiusa.org","200+ AMI schools in US. School Locator tool."],
["AMS (American Montessori Society)","USA","amshq.org","amshq.org","Largest US Montessori org."],
["Montessori Australia","Australia","+61 2 9986 2282","montessori.org.au","National peak body since 2007."],
["AMI Canada","Canada","ami-canada.com","ami-canada.com","MQA program since 1980s."],
["Foundation for Montessori Education","Toronto","416-769-7457","montessori-ami.ca","AMI training centre."],
["Indian Montessori Foundation","India","montessori-india.org","montessori-india.org","Flagship schools directory with contacts."],
["Montessori Mexico","Mexico","montessorimx.com","montessorimx.com","AMI affiliated. Annual congress."],
["Montessori Europe","Europe","office@montessori-europe.net","montessori-europe.net","Pan-European network."],
["Deutsche Montessori Gesellschaft","Berlin","Founded 1925 with Maria Montessori","dmg-montessori.de","AMI affiliated Germany."],
["South African Montessori Assoc","South Africa","samontessori.org.za","samontessori.org.za","Find a School directory."],
["Montessori Society AMI (UK)","UK","montessorisociety.org.uk","montessorisociety.org.uk","UK AMI affiliate."],
]
for r in orgs:
    ws3.append(r)
sh(ws3, 5)
sd(ws3, 13, 5)
for i, w in enumerate([40,20,40,30,55], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = 'A2'

# ===== SHEET 4: COLD EMAIL =====
ws4 = wb.create_sheet("Cold Email Template")
ws4.column_dimensions['A'].width = 80
ws4['A1'] = "MONTREE COLD EMAIL TEMPLATE"
ws4['A1'].font = Font(name='Arial', bold=True, size=14, color='0D3330')
ws4['A3'] = "Subject: A quick look at something new for Montessori"
ws4['A3'].font = Font(name='Arial', bold=True, size=11)
ws4['A5'] = "Dear [Name],"
ws4['A7'] = "I built Montree for my own classroom in Beijing — a small tool to help Montessori teachers spend less time on admin and more time with children."
ws4['A9'] = "The core idea is simple: take a photo of a child working. The AI identifies the Montessori material, updates the child's progress records, and prepares a parent update — all automatically."
ws4['A11'] = "It knows 329 works across the full AMI curriculum. It has 13 developmental psychologists built in. And if it ever gets something wrong, correct it once — it learns and never makes that mistake again."
ws4['A13'] = "No other Montessori software does any of this. We checked."
ws4['A15'] = "Your school is already excellent. This is how you stay ahead — not just keeping up, but setting the pace."
ws4['A17'] = "Let me know if you would like more information and it would be my pleasure to answer any questions you may have."
ws4['A19'] = "Warm regards,"
ws4['A20'] = "Tredoux"
ws4['A21'] = "Montree — montree.xyz"
for r in [5,7,9,11,13,15,17,19,20,21]:
    ws4[f'A{r}'].font = Font(name='Arial', size=11)
    ws4[f'A{r}'].alignment = Alignment(wrap_text=True)

# ===== SAVE =====
out = "/Users/tredouxwillemse/Desktop/Master Brain/ACTIVE/whale/docs/Montree_Global_Outreach_List.xlsx"
wb.save(out)
print(f"Saved to {out}")
print("Done! Sheets: Top 50, Chains, Training & Orgs, Cold Email Template")
